"""
Excel Handler for London Property Search Analyzer
Handles Excel file processing, data import/export, and formatting
"""

import pandas as pd
import numpy as np
from datetime import datetime
import io
from typing import Dict, List, Optional, Tuple, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
import xlsxwriter


class ExcelHandler:
    """
    Handles Excel file operations for property data
    """

    def __init__(self):
        self.default_columns = [
            'address', 'property_type', 'bedrooms', 'bathrooms', 
            'price', 'square_feet', 'price_per_sqft', 'borough',
            'features', 'days_on_market', 'listing_date', 'agent'
        ]

    def export_properties_to_excel(self, properties: List[Dict], 
                                 filename: Optional[str] = None) -> io.BytesIO:
        """
        Export property data to Excel with formatting and charts

        Args:
            properties: List of property dictionaries
            filename: Optional filename for saving to disk

        Returns:
            BytesIO object containing Excel file
        """
        if not properties:
            raise ValueError("No properties to export")

        # Create DataFrame
        df = pd.DataFrame(properties)

        # Create BytesIO object
        output = io.BytesIO()

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main properties sheet
            df.to_excel(writer, sheet_name='Properties', index=False)

            # Summary statistics sheet
            self._create_summary_sheet(writer, df)

            # Borough analysis sheet
            self._create_borough_analysis_sheet(writer, df)

            # Property type analysis sheet
            self._create_property_type_sheet(writer, df)

            # Charts sheet
            self._create_charts_sheet(writer, df)

        # Apply formatting
        output.seek(0)
        self._apply_formatting(output)

        # Save to file if filename provided
        if filename:
            with open(filename, 'wb') as f:
                f.write(output.getvalue())

        output.seek(0)
        return output

    def _create_summary_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Create summary statistics sheet"""

        # Calculate summary statistics
        summary_data = {
            'Metric': [
                'Total Properties',
                'Average Price',
                'Median Price',
                'Min Price',
                'Max Price',
                'Price Standard Deviation',
                'Average Size (sq ft)',
                'Average Price per Sq Ft',
                'Most Common Property Type',
                'Most Common Borough',
                'Properties with Garden',
                'Properties with Parking',
                'Average Days on Market'
            ],
            'Value': [
                len(df),
                f"£{df['price'].mean():,.0f}",
                f"£{df['price'].median():,.0f}",
                f"£{df['price'].min():,.0f}",
                f"£{df['price'].max():,.0f}",
                f"£{df['price'].std():,.0f}",
                f"{df['square_feet'].mean():.0f}",
                f"£{df['price_per_sqft'].mean():.0f}",
                df['property_type'].mode().iloc[0],
                df['borough'].mode().iloc[0],
                len(df[df['features'].astype(str).str.contains('Garden', na=False)]),
                len(df[df['features'].astype(str).str.contains('Parking', na=False)]),
                f"{df['days_on_market'].mean():.0f}"
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    def _create_borough_analysis_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Create borough analysis sheet"""

        borough_stats = df.groupby('borough').agg({
            'price': ['count', 'mean', 'median', 'std'],
            'square_feet': 'mean',
            'price_per_sqft': 'mean',
            'days_on_market': 'mean'
        }).round(2)

        # Flatten column names
        borough_stats.columns = [
            'Property Count', 'Average Price', 'Median Price', 'Price Std Dev',
            'Avg Size (sq ft)', 'Avg Price per Sq Ft', 'Avg Days on Market'
        ]

        borough_stats.to_excel(writer, sheet_name='Borough Analysis')

    def _create_property_type_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Create property type analysis sheet"""

        type_stats = df.groupby('property_type').agg({
            'price': ['count', 'mean', 'median'],
            'square_feet': 'mean',
            'price_per_sqft': 'mean',
            'bedrooms': 'mean'
        }).round(2)

        # Flatten column names
        type_stats.columns = [
            'Count', 'Average Price', 'Median Price',
            'Avg Size (sq ft)', 'Avg Price per Sq Ft', 'Avg Bedrooms'
        ]

        type_stats.to_excel(writer, sheet_name='Property Types')

    def _create_charts_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Create charts sheet with summary visualizations"""

        # Create data for charts
        borough_summary = df.groupby('borough')['price'].mean().reset_index()
        borough_summary.columns = ['Borough', 'Average Price']

        type_summary = df['property_type'].value_counts().reset_index()
        type_summary.columns = ['Property Type', 'Count']

        # Write chart data
        borough_summary.to_excel(writer, sheet_name='Chart Data', 
                                startrow=0, startcol=0, index=False)

        type_summary.to_excel(writer, sheet_name='Chart Data', 
                             startrow=0, startcol=4, index=False)

    def _apply_formatting(self, excel_file: io.BytesIO):
        """Apply advanced formatting to the Excel file"""

        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Format Properties sheet
        if 'Properties' in workbook.sheetnames:
            ws = workbook['Properties']

            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Apply number formatting to price columns
            for row in ws.iter_rows(min_row=2):
                if row[4].value:  # Price column
                    row[4].number_format = '£#,##0'
                if row[6].value:  # Price per sq ft column
                    row[6].number_format = '£#,##0'

        # Format Summary sheet
        if 'Summary' in workbook.sheetnames:
            ws = workbook['Summary']

            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Highlight key metrics
            highlight_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            key_metrics = [1, 2, 3]  # Total, Average, Median rows

            for row_num in key_metrics:
                for cell in ws[row_num + 1]:  # +1 because of header
                    cell.fill = highlight_fill

        # Save the formatted workbook
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def import_properties_from_excel(self, file_path: str) -> List[Dict]:
        """
        Import property data from Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            List of property dictionaries
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path)

            # Validate required columns
            required_columns = ['address', 'property_type', 'price']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")

            # Clean and validate data
            df = self._clean_imported_data(df)

            # Convert to list of dictionaries
            properties = df.to_dict('records')

            return properties

        except Exception as e:
            raise Exception(f"Error importing Excel file: {str(e)}")

    def _clean_imported_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate imported property data"""

        # Remove rows with missing essential data
        df = df.dropna(subset=['address', 'property_type', 'price'])

        # Clean price data
        if df['price'].dtype == 'object':
            # Remove currency symbols and convert to numeric
            df['price'] = df['price'].astype(str).str.replace('[£,]', '', regex=True)
            df['price'] = pd.to_numeric(df['price'], errors='coerce')

        # Validate price range
        df = df[(df['price'] >= 50000) & (df['price'] <= 50000000)]

        # Clean bedrooms data
        if 'bedrooms' in df.columns:
            df['bedrooms'] = pd.to_numeric(df['bedrooms'], errors='coerce')
            df['bedrooms'] = df['bedrooms'].fillna(0).astype(int)

        # Clean square feet data
        if 'square_feet' in df.columns:
            df['square_feet'] = pd.to_numeric(df['square_feet'], errors='coerce')
            # Calculate price per sq ft if missing
            if 'price_per_sqft' not in df.columns:
                df['price_per_sqft'] = df['price'] / df['square_feet']

        # Standardize property types
        property_type_mapping = {
            'apartment': 'Flat',
            'flat': 'Flat',
            'house': 'House',
            'studio': 'Studio',
            'penthouse': 'Penthouse',
            'maisonette': 'Maisonette'
        }

        if 'property_type' in df.columns:
            df['property_type'] = df['property_type'].str.lower().map(
                lambda x: property_type_mapping.get(x, x.title())
            )

        return df

    def create_template_file(self, template_type: str = 'basic') -> io.BytesIO:
        """
        Create Excel template for property data import

        Args:
            template_type: 'basic' or 'advanced' template

        Returns:
            BytesIO object containing template file
        """

        if template_type == 'basic':
            columns = [
                'address', 'property_type', 'bedrooms', 'bathrooms',
                'price', 'square_feet', 'borough'
            ]
            sample_data = [{
                'address': '123 Example Street, London',
                'property_type': 'Flat',
                'bedrooms': 2,
                'bathrooms': 1,
                'price': 500000,
                'square_feet': 800,
                'borough': 'Camden'
            }]

        else:  # advanced
            columns = [
                'address', 'property_type', 'bedrooms', 'bathrooms',
                'price', 'square_feet', 'borough', 'features',
                'listing_date', 'agent_name', 'agent_phone',
                'energy_rating', 'council_tax_band'
            ]
            sample_data = [{
                'address': '123 Example Street, London',
                'property_type': 'Flat',
                'bedrooms': 2,
                'bathrooms': 1,
                'price': 500000,
                'square_feet': 800,
                'borough': 'Camden',
                'features': 'Garden, Parking, Balcony',
                'listing_date': datetime.now().strftime('%Y-%m-%d'),
                'agent_name': 'Example Estate Agents',
                'agent_phone': '020 1234 5678',
                'energy_rating': 'C',
                'council_tax_band': 'D'
            }]

        # Create DataFrame
        df = pd.DataFrame([{col: '' for col in columns}] + sample_data)

        # Create Excel file
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Property Data', index=False)

            # Add instructions sheet
            instructions = pd.DataFrame({
                'Column': columns,
                'Description': self._get_column_descriptions(columns),
                'Required': ['Yes' if col in ['address', 'property_type', 'price'] else 'No' for col in columns]
            })

            instructions.to_excel(writer, sheet_name='Instructions', index=False)

        output.seek(0)
        return output

    def _get_column_descriptions(self, columns: List[str]) -> List[str]:
        """Get descriptions for template columns"""

        descriptions = {
            'address': 'Full property address including postcode',
            'property_type': 'Type: Flat, House, Studio, Penthouse, Maisonette',
            'bedrooms': 'Number of bedrooms (numeric)',
            'bathrooms': 'Number of bathrooms (numeric)',
            'price': 'Property price in GBP (numeric, no symbols)',
            'square_feet': 'Property size in square feet (numeric)',
            'borough': 'London borough name',
            'features': 'Comma-separated features (Garden, Parking, etc.)',
            'listing_date': 'Date listed (YYYY-MM-DD format)',
            'agent_name': 'Estate agent name',
            'agent_phone': 'Agent contact phone',
            'energy_rating': 'Energy efficiency rating (A-G)',
            'council_tax_band': 'Council tax band (A-H)'
        }

        return [descriptions.get(col, 'Additional property information') for col in columns]

    def validate_excel_data(self, file_path: str) -> Dict:
        """
        Validate Excel file data and return validation report

        Args:
            file_path: Path to Excel file

        Returns:
            Validation report dictionary
        """

        try:
            df = pd.read_excel(file_path)

            validation_report = {
                'valid': True,
                'errors': [],
                'warnings': [],
                'statistics': {
                    'total_rows': len(df),
                    'valid_rows': 0,
                    'invalid_rows': 0
                }
            }

            # Check required columns
            required_columns = ['address', 'property_type', 'price']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                validation_report['valid'] = False
                validation_report['errors'].append(f"Missing required columns: {missing_columns}")
                return validation_report

            # Validate each row
            valid_rows = 0

            for index, row in df.iterrows():
                row_errors = []

                # Validate address
                if pd.isna(row['address']) or str(row['address']).strip() == '':
                    row_errors.append(f"Row {index + 2}: Missing address")

                # Validate property type
                valid_types = ['Flat', 'House', 'Studio', 'Penthouse', 'Maisonette']
                if row['property_type'] not in valid_types:
                    row_errors.append(f"Row {index + 2}: Invalid property type")

                # Validate price
                try:
                    price = float(str(row['price']).replace('£', '').replace(',', ''))
                    if price <= 0 or price > 50000000:
                        row_errors.append(f"Row {index + 2}: Invalid price range")
                except:
                    row_errors.append(f"Row {index + 2}: Invalid price format")

                # Validate bedrooms if present
                if 'bedrooms' in df.columns and not pd.isna(row['bedrooms']):
                    try:
                        bedrooms = int(row['bedrooms'])
                        if bedrooms < 0 or bedrooms > 10:
                            row_errors.append(f"Row {index + 2}: Invalid bedroom count")
                    except:
                        row_errors.append(f"Row {index + 2}: Invalid bedroom format")

                if row_errors:
                    validation_report['errors'].extend(row_errors)
                else:
                    valid_rows += 1

            validation_report['statistics']['valid_rows'] = valid_rows
            validation_report['statistics']['invalid_rows'] = len(df) - valid_rows

            if validation_report['errors']:
                validation_report['valid'] = False

            # Add warnings for missing optional columns
            optional_columns = ['bedrooms', 'bathrooms', 'square_feet', 'borough']
            for col in optional_columns:
                if col not in df.columns:
                    validation_report['warnings'].append(f"Optional column '{col}' not found")

            return validation_report

        except Exception as e:
            return {
                'valid': False,
                'errors': [f"Error reading file: {str(e)}"],
                'warnings': [],
                'statistics': {'total_rows': 0, 'valid_rows': 0, 'invalid_rows': 0}
            }


# Utility functions
def create_sample_excel():
    """Create sample Excel file for testing"""
    handler = ExcelHandler()

    # Sample property data
    properties = [
        {
            'address': '123 Kings Road, Chelsea',
            'property_type': 'Flat',
            'bedrooms': 2,
            'bathrooms': 2,
            'price': 850000,
            'square_feet': 900,
            'price_per_sqft': 944.44,
            'borough': 'Kensington and Chelsea',
            'features': ['Garden', 'Parking'],
            'days_on_market': 45,
            'listing_date': '2023-09-01',
            'agent': {'name': 'Foxtons', 'phone': '020 7000 1234'}
        }
    ]

    return handler.export_properties_to_excel(properties)
